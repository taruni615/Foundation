import os
import re
import pymysql
import pandas as pd
from topicwise_pipeline import DB_HOST, DB_NAME, DB_PASSWORD, DB_PORT, DB_USER

# Copy functions from bank_read.py to derive attributes and estimate difficulty
SUBJECTS = ["Physics", "Chemistry", "Biology", "Mathematics", "Science"]
_CLASS_RE = re.compile(r"class\s*(\d{1,2})")
_CLASS_TH_RE = re.compile(r"(\d{1,2})\s*th")
_CLASS_LEAD_RE = re.compile(r"^\s*(\d{1,2})\b")

# Regex to match characters not allowed in XML (Excel xlsx files)
# XML 1.0 permits: #x9 | #xA | #xD | [#x20-#xD7FF] | [#xE000-#xFFFD] | [#x10000-#x10FFFF]
# So we filter out control characters except tab (\t), line feed (\n), and carriage return (\r).
ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]')

def clean_string(val):
    if isinstance(val, str):
        return ILLEGAL_CHARACTERS_RE.sub('', val)
    return val

def derive_attributes(book_slug: str):
    low = (book_slug or "").lower()
    subject = ""
    for s in SUBJECTS:
        if s.lower() in low:
            subject = s
            break
    if not subject and "math" in low:
        subject = "Mathematics"
    cls = ""
    m = _CLASS_RE.search(low) or _CLASS_TH_RE.search(low) or _CLASS_LEAD_RE.match(low)
    if m:
        cls = m.group(1)
    board = "Foundation" if "foundation" in low else ""
    return subject, cls, board

def estimate_difficulty(stem: str, question_type: str = "") -> str:
    s = stem or ""
    qt = (question_type or "").lower()
    score = 0
    n = len(s)
    if n > 260:
        score += 2
    elif n > 130:
        score += 1
    if "<math" in s or "$" in s or "\\(" in s or "\\frac" in s:
        score += 1
    if any(t in qt for t in ("numerical", "problem", "hots", "assertion", "matrix")):
        score += 1
    if any(t in qt for t in ("definition", "fill", "true", "one word", "short")):
        score -= 1
    if score >= 3:
        return "Difficult"
    if score >= 1:
        return "Moderate"
    return "Easy"

def main():
    print("Connecting to MySQL...")
    conn = pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor
    )
    
    try:
        print("Querying all questions...")
        query = """
        SELECT 
            r.id,
            r.book_slug,
            c.chapter_number,
            r.chapter_name,
            r.question_type,
            r.question,
            r.answer
        FROM qa_content_row r
        LEFT JOIN qa_chapter c ON r.chapter_id = c.chapter_id
        ORDER BY r.book_slug, c.chapter_number, r.id
        """
        
        with conn.cursor() as cursor:
            cursor.execute(query)
            rows = cursor.fetchall()
            
        print(f"Fetched {len(rows)} questions. Processing and cleaning attributes...")
        
        processed_data = []
        for row in rows:
            book_slug = row['book_slug']
            subject, cls, board = derive_attributes(book_slug)
            difficulty = estimate_difficulty(row['question'], row['question_type'])
            
            # Clean all string values to prevent Excel generation errors
            q_text = clean_string(row['question'])
            a_text = clean_string(row['answer'])
            ch_name = clean_string(row['chapter_name'])
            book_slug_cleaned = clean_string(book_slug)
            q_type = clean_string(row['question_type'])
            
            processed_data.append({
                "Question ID": row['id'],
                "Book": book_slug_cleaned,
                "Subject": subject,
                "Class": f"Class {cls}" if cls else "",
                "Board/Program": board,
                "Chapter Number": row['chapter_number'] if row['chapter_number'] is not None else "",
                "Chapter Name": ch_name,
                "Question Type": q_type,
                "Difficulty": difficulty,
                "Question": q_text,
                "Answer": a_text
            })
            
        df = pd.DataFrame(processed_data)
        
        output_file = "all_database_questions.xlsx"
        print(f"Saving {len(df)} rows to {output_file}...")
        
        # Write to Excel. We can format it nicely using openpyxl.
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Questions", index=False)
            
            # Format sheet columns width and styling
            workbook = writer.book
            worksheet = writer.sheets["Questions"]
            
            # Style header
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                
            # Adjust row height and alignment
            worksheet.row_dimensions[1].height = 28
            
            # Set specific column widths
            col_widths = {
                "Question ID": 12,
                "Book": 25,
                "Subject": 15,
                "Class": 12,
                "Board/Program": 15,
                "Chapter Number": 15,
                "Chapter Name": 25,
                "Question Type": 20,
                "Difficulty": 12,
                "Question": 60,
                "Answer": 60
            }
            
            for col_idx, col_name in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                width = col_widths.get(col_name, 15)
                worksheet.column_dimensions[col_letter].width = width
                
                # Apply left alignment with word wrap for text columns
                if col_name in ["Question", "Answer"]:
                    for row_idx in range(2, len(df) + 2):
                        worksheet.cell(row=row_idx, column=col_idx).alignment = align_left
                else:
                    for row_idx in range(2, len(df) + 2):
                        worksheet.cell(row=row_idx, column=col_idx).alignment = align_center
                        
        print("Excel file created successfully!")
        
    finally:
        conn.close()

if __name__ == "__main__":
    main()
